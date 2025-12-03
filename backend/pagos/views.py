from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.filters import SearchFilter, OrderingFilter
from django_filters.rest_framework import DjangoFilterBackend
from django.utils import timezone
from .models import PagoPersona, ComprobantePago, PagoComprobante, PagoCambioPersona, Prepago, PagoProveedor
from .serializers import (
	PagoPersonaSerializer,
	ComprobantePagoSerializer,
	PagoComprobanteSerializer,
	PagoCambioPersonaSerializer,
	PrepagoSerializer,
    PagoProveedorSerializer,
)
from personas.models import Persona, PersonaCurso, PersonaGrupo
from cursos.models import Curso
from decimal import Decimal
import csv
import io
from django.http import HttpResponse



class PagoProveedorViewSet(viewsets.ModelViewSet):
    queryset = PagoProveedor.objects.all().select_related('prv_id', 'coc_id', 'usu_id')
    serializer_class = PagoProveedorSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ['prv_id', 'coc_id', 'usu_id']
    search_fields = ['prv_id__prv_nombre_fantasia', 'prv_id__prv_rut', 'ppr_observacion']
    ordering_fields = ['ppr_fecha', 'ppr_valor']

class PagoPersonaViewSet(viewsets.ModelViewSet):
	queryset = PagoPersona.objects.all().select_related('per_id', 'cur_id')
	serializer_class = PagoPersonaSerializer
	permission_classes = [IsAuthenticated]
	filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
	filterset_fields = ['per_id', 'cur_id', 'usu_id', 'pap_tipo', 'pap_estado']
	search_fields = ['per_id__per_run', 'per_id__per_nombres', 'per_id__per_apelpat', 'per_id__per_apelmat']
	ordering_fields = ['pap_fecha_hora', 'pap_valor']

	def create(self, request, *args, **kwargs):
		serializer = self.get_serializer(data=request.data)
		serializer.is_valid(raise_exception=True)
		
		# Extract extra fields
		file = request.FILES.get('file')
		coc_id = request.data.get('coc_id')
		
		# Perform standard create
		self.perform_create(serializer)
		payment = serializer.instance
		
		# If file and concept are provided, create ComprobantePago
		if file and coc_id:
			try:
				from maestros.models import ConceptoContable
				concepto = ConceptoContable.objects.get(coc_id=coc_id)
				
				# Determine pec_id (PersonaCurso) - try to find one active
				pec_id = None
				persona_curso = PersonaCurso.objects.filter(per_id=payment.per_id, cus_id__cur_id=payment.cur_id).first()
				if persona_curso:
					pec_id = persona_curso
				else:
					# Fallback: any active enrollment
					pec_id = PersonaCurso.objects.filter(per_id=payment.per_id).first()
				
				if pec_id:
					comprobante = ComprobantePago.objects.create(
						usu_id=payment.usu_id,
						pec_id=pec_id,
						coc_id=concepto,
						cpa_fecha_hora=timezone.now(),
						cpa_fecha=timezone.now().date(),
						cpa_numero=0, # Placeholder, logic for numbering can be added
						cpa_valor=payment.pap_valor,
						cpa_archivo=file,
						cpa_tipo=ComprobantePago.CPA_TIPO_INGRESO if payment.pap_tipo == PagoPersona.PAP_TIPO_INGRESO else ComprobantePago.CPA_TIPO_EGRESO
					)
					
					# Link payment to comprobante
					PagoComprobante.objects.create(pap_id=payment, cpa_id=comprobante)
			except Exception as e:
				print(f"Error creating automatic comprobante: {e}")
				# We don't fail the request if comprobante creation fails, but we log it
		
		headers = self.get_success_headers(serializer.data)
		return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)

	@action(detail=False, methods=['post'], url_path='masivo')
	def registro_masivo(self, request):
		"""Registro masivo de pagos: valor_total se reparte entre personas seleccionadas o filtradas por curso/grupo/rama.
		Payload: {"valor_total": 100000, "valor_unitario": 5000, "personas": [1,2], "cur_id": 1, "gru_id": 1, "ram_id": 1, "pap_tipo": 1}
		Supports multipart/form-data for file upload.
		"""
		valor_total = request.data.get('valor_total')
		valor_unitario = request.data.get('valor_unitario')
		
		# Handle personas list (support both JSON list and FormData getlist)
		personas = request.data.get('personas')
		if not isinstance(personas, list):
			personas = request.data.getlist('personas') if hasattr(request.data, 'getlist') else []
			# If it's a list of strings/ints, ensure they are clean
			if personas and isinstance(personas[0], str) and ',' in personas[0]:
				# Handle case where FormData sends "1,2,3" as a single string
				personas = personas[0].split(',')

		cur_id = request.data.get('cur_id')
		gru_id = request.data.get('gru_id')
		ram_id = request.data.get('ram_id')
		pap_tipo = int(request.data.get('pap_tipo', PagoPersona.PAP_TIPO_INGRESO))
		
		# Extra fields for Comprobante
		file = request.FILES.get('file')
		coc_id = request.data.get('coc_id')

		if not valor_total and not valor_unitario:
			return Response({'detail': 'Debe proporcionar valor_total o valor_unitario.'}, status=status.HTTP_400_BAD_REQUEST)

		# Obtain list of persona instances to create payments for
		queryset = Persona.objects.all()
		if personas:
			queryset = queryset.filter(per_id__in=personas)
		if cur_id:
			queryset = queryset.filter(personacurso__cus_id__cur_id=cur_id)  # Personas en el curso
		if gru_id:
			queryset = queryset.filter(personagrupo__gru_id=gru_id)  # Personas en el grupo
		if ram_id:
			# Filtrar por Rama a través de la sección del curso
			queryset = queryset.filter(personacurso__cus_id__ram_id=ram_id)

		queryset = queryset.distinct()
		total_personas = queryset.count()
		if total_personas == 0:
			return Response({'detail': 'No hay personas que coincidan con los filtros.'}, status=status.HTTP_400_BAD_REQUEST)

		if valor_unitario:
			valor_por_persona = Decimal(str(valor_unitario))
			valor_total_dec = valor_por_persona * total_personas
		else:
			valor_total_dec = Decimal(str(valor_total))
			valor_por_persona = (valor_total_dec / total_personas).quantize(Decimal('0.01'))

		created = []
		
		# Prepare Comprobante if applicable
		comprobante = None
		if file and coc_id:
			try:
				from maestros.models import ConceptoContable
				concepto = ConceptoContable.objects.get(coc_id=coc_id)
				
				# Find a representative pec_id (use the first person's enrollment in this course)
				first_persona = queryset.first()
				pec_id = None
				if cur_id:
					pec_id = PersonaCurso.objects.filter(per_id=first_persona, cus_id__cur_id=cur_id).first()
				else:
					pec_id = PersonaCurso.objects.filter(per_id=first_persona).first()
				
				if pec_id:
					# Determine user
					usu_comprobante = None
					if hasattr(request.user, 'usuario'):
						usu_comprobante = request.user.usuario
					elif request.data.get('usu_id'):
						from usuarios.models import Usuario as UsuarioModel
						try:
							usu_comprobante = UsuarioModel.objects.get(usu_id=request.data.get('usu_id'))
						except:
							pass

					if usu_comprobante:
						comprobante = ComprobantePago.objects.create(
							usu_id=usu_comprobante,
							pec_id=pec_id,
							coc_id=concepto,
							cpa_fecha_hora=timezone.now(),
							cpa_fecha=timezone.now().date(),
							cpa_numero=0,
							cpa_valor=valor_total_dec, # Total value for the mass payment
							cpa_archivo=file,
							cpa_tipo=ComprobantePago.CPA_TIPO_INGRESO if pap_tipo == PagoPersona.PAP_TIPO_INGRESO else ComprobantePago.CPA_TIPO_EGRESO
						)
			except Exception as e:
				print(f"Error creating mass comprobante: {e}")

		for persona in queryset:
			curso_instance = None
			if cur_id:
				try:
					curso_instance = Curso.objects.get(cur_id=cur_id)
				except Curso.DoesNotExist:
					return Response({'detail': 'Curso no encontrado.'}, status=status.HTTP_404_NOT_FOUND)
			else:
				# try to find the course from PersonaCurso
				persona_curso = PersonaCurso.objects.filter(per_id=persona).first()
				if persona_curso:
					curso_instance = persona_curso.cus_id.cur_id

			# Determine usuario: prefer Usuario linked to request.user, if available
			usu = None
			if hasattr(request.user, 'usuario'):
				usu = request.user.usuario
			else:
				from usuarios.models import Usuario as UsuarioModel
				if getattr(request.user, 'username', None):
					usu = UsuarioModel.objects.filter(usu_username=request.user.username).first()
				if not usu and getattr(request.user, 'email', None):
					usu = UsuarioModel.objects.filter(usu_email=request.user.email).first()
				# fallback to the provided 'usu_id' in request data
				if not usu and request.data.get('usu_id'):
					try:
						usu = UsuarioModel.objects.get(usu_id=request.data.get('usu_id'))
					except UsuarioModel.DoesNotExist:
						pass

			if not usu:
				return Response({'detail': 'No se pudo determinar el Usuario (usu_id). Proporcione usu_id en la petición o asocie un Usuario válido.'}, status=status.HTTP_400_BAD_REQUEST)

			payment = PagoPersona.objects.create(
				per_id=persona,
				cur_id=curso_instance,
				usu_id=usu if usu else None,
				pap_fecha_hora=timezone.now(),
				pap_tipo=pap_tipo,
				pap_valor=valor_por_persona,
				pap_observacion=f'Registro masivo: {valor_total_dec} (valor por persona: {valor_por_persona})'
			)
			created.append(payment)
			
			# Link to comprobante if created
			if comprobante:
				PagoComprobante.objects.create(pap_id=payment, cpa_id=comprobante)

		serializer = self.get_serializer(created, many=True)
		return Response({'created_count': len(created), 'valor_por_persona': str(valor_por_persona), 'pagos': serializer.data}, status=status.HTTP_201_CREATED)

	@action(detail=False, methods=['post'], url_path='multi-persona')
	def registro_multi_persona(self, request):
		"""Registro de pagos donde un pagador (payer_id) paga por varias personas (beneficiaries).
		Payload: {
			"payer_id": 1, 
			"cur_id": 1, 
			"payments": [
				{"per_id": 2, "amount": 5000},
				{"per_id": 3, "amount": 5000}
			],
			"pap_tipo": 1,
			"observacion": "Pago cuota enero"
		}
		"""
		payer_id = request.data.get('payer_id') # Persona que paga (opcional, solo para registro)
		cur_id = request.data.get('cur_id')
		payments = request.data.get('payments', [])
		pap_tipo = int(request.data.get('pap_tipo', PagoPersona.PAP_TIPO_INGRESO))
		observacion = request.data.get('observacion', '')

		if not payments:
			return Response({'detail': 'Lista de pagos (payments) requerida.'}, status=status.HTTP_400_BAD_REQUEST)

		# Determine usuario
		usu = None
		if hasattr(request.user, 'usuario'):
			usu = request.user.usuario
		else:
			from usuarios.models import Usuario as UsuarioModel
			if getattr(request.user, 'username', None):
				usu = UsuarioModel.objects.filter(usu_username=request.user.username).first()
			if not usu and getattr(request.user, 'email', None):
				usu = UsuarioModel.objects.filter(usu_email=request.user.email).first()
		
		if not usu:
			# Try to get from request if not authenticated properly
			if request.data.get('usu_id'):
				try:
					usu = UsuarioModel.objects.get(usu_id=request.data.get('usu_id'))
				except:
					pass
		
		if not usu:
			return Response({'detail': 'Usuario no identificado.'}, status=status.HTTP_400_BAD_REQUEST)

		created = []
		for item in payments:
			per_id = item.get('per_id')
			amount = item.get('amount')
			
			if not per_id or not amount:
				continue

			try:
				persona = Persona.objects.get(per_id=per_id)
			except Persona.DoesNotExist:
				continue

			curso_instance = None
			if cur_id:
				try:
					curso_instance = Curso.objects.get(cur_id=cur_id)
				except Curso.DoesNotExist:
					pass
			
			if not curso_instance:
				# Try to infer course
				persona_curso = PersonaCurso.objects.filter(per_id=persona).first()
				if persona_curso:
					curso_instance = persona_curso.cus_id.cur_id
			
			if not curso_instance:
				continue # Skip if no course found

			payment = PagoPersona.objects.create(
				per_id=persona,
				cur_id=curso_instance,
				usu_id=usu,
				pap_fecha_hora=timezone.now(),
				pap_tipo=pap_tipo,
				pap_valor=amount,
				pap_observacion=f"{observacion} (Pagado por ID: {payer_id})" if payer_id else observacion
			)
			created.append(payment)

		serializer = self.get_serializer(created, many=True)
		return Response({'created_count': len(created), 'pagos': serializer.data}, status=status.HTTP_201_CREATED)

	@action(detail=False, methods=['get'], url_path='estado-cuenta')
	def estado_cuenta(self, request):
		"""Obtiene el estado de cuenta de una persona en un curso (todos los ingresos y egresos).
		Params: cur_id, per_id
		"""
		cur_id = request.query_params.get('cur_id')
		per_id = request.query_params.get('per_id')

		if not cur_id or not per_id:
			return Response({'detail': 'cur_id y per_id son requeridos.'}, status=status.HTTP_400_BAD_REQUEST)

		pagos = PagoPersona.objects.filter(cur_id=cur_id, per_id=per_id).order_by('pap_fecha_hora')
		
		total_ingresos = sum(p.pap_valor for p in pagos if p.pap_tipo == PagoPersona.PAP_TIPO_INGRESO)
		total_egresos = sum(p.pap_valor for p in pagos if p.pap_tipo == PagoPersona.PAP_TIPO_EGRESO)
		balance = total_ingresos - total_egresos

		serializer = self.get_serializer(pagos, many=True)
		
		return Response({
			'resumen': {
				'total_ingresos': total_ingresos,
				'total_egresos': total_egresos,
				'balance': balance
			},
			'transacciones': serializer.data
		})

	@action(detail=False, methods=['get'], url_path='dashboard')
	def dashboard(self, request):
		"""Dashboard financiero del curso.
		Params: cur_id
		"""
		cur_id = request.query_params.get('cur_id')
		if not cur_id:
			return Response({'detail': 'cur_id es requerido.'}, status=status.HTTP_400_BAD_REQUEST)

		pagos = PagoPersona.objects.filter(cur_id=cur_id)
		
		total_ingresos = sum(p.pap_valor for p in pagos if p.pap_tipo == PagoPersona.PAP_TIPO_INGRESO)
		total_egresos = sum(p.pap_valor for p in pagos if p.pap_tipo == PagoPersona.PAP_TIPO_EGRESO)
		balance = total_ingresos - total_egresos

		# Agrupar por fecha (mes)
		from django.db.models.functions import TruncMonth
		from django.db.models import Sum
		
		por_mes = pagos.annotate(month=TruncMonth('pap_fecha_hora')).values('month', 'pap_tipo').annotate(total=Sum('pap_valor')).order_by('month')

		return Response({
			'global': {
				'total_ingresos': total_ingresos,
				'total_egresos': total_egresos,
				'balance': balance
			},
			'evolucion': por_mes
		})

	@action(detail=True, methods=['post'], url_path='cambio-persona')
	def cambio_persona(self, request, pk=None):
		"""Cambio de persona en un pago: registra la modificación en PagoCambioPersona y actualiza PagoPersona.per_id.
		Payload: {"new_per_id": 123}
		"""
		payment = self.get_object()
		new_per_id = request.data.get('new_per_id')
		if not new_per_id:
			return Response({'detail': 'new_per_id es requerido.'}, status=status.HTTP_400_BAD_REQUEST)

		try:
			new_persona = Persona.objects.get(per_id=new_per_id)
		except Persona.DoesNotExist:
			return Response({'detail': 'Persona no encontrada.'}, status=status.HTTP_404_NOT_FOUND)

		# Log the change
		PagoCambioPersona.objects.create(per_id=payment.per_id, pap_id=payment, usu_id=request.user.usuario if hasattr(request.user, 'usuario') else request.user, pcp_fecha_hora=timezone.now())

		# Update payment
		payment.per_id = new_persona
		payment.save()

		serializer = self.get_serializer(payment)
		return Response(serializer.data, status=status.HTTP_200_OK)

	@action(detail=False, methods=['get'], url_path='export')
	def export(self, request):
		"""Export payments to CSV (readable in Excel) or Excel xlsx (optionally) using query params csv=1 or xlsx=1"""
		per_id = request.query_params.get('per_id')
		cur_id = request.query_params.get('cur_id')
		format = 'xlsx' if request.query_params.get('xlsx') else 'csv'

		qs = self.filter_queryset(self.get_queryset())
		if per_id:
			qs = qs.filter(per_id__per_id=per_id)
		if cur_id:
			qs = qs.filter(cur_id__cur_id=cur_id)

		# Prepare CSV
		if format == 'csv':
			response = HttpResponse(content_type='text/csv')
			response['Content-Disposition'] = 'attachment; filename="pagos_export.csv"'
			writer = csv.writer(response)
			headers = ['Rut', 'Nombre', 'Apellidos', 'Correo electrónico', 'Grupo', 'Distrito', 'Zona', 'Curso', 'Valor Cuota', 'Valor pagado', 'Valor Adeudado', 'Fecha ultimo pago']
			writer.writerow(headers)
			for p in qs.select_related('per_id', 'cur_id'):
				persona = p.per_id
				# compute fields
				curso = getattr(p.cur_id, 'cur_id', p.cur_id) if p.cur_id else None
				valor_cuota = curso.cur_cuota_con_almuerzo if curso and hasattr(curso, 'cur_cuota_con_almuerzo') else ''
				valor_pagado = p.pap_valor
				# valor adeudado: cuota - pagado
				if valor_cuota and valor_pagado:
					valor_adeudado = Decimal(valor_cuota) - Decimal(valor_pagado)
				else:
					valor_adeudado = ''
				fecha_ultimo_pago = p.pap_fecha_hora
				grupo = None
				persona_grupo = PersonaGrupo.objects.filter(per_id=persona).first()
				if persona_grupo:
					grupo = persona_grupo.gru_id.gru_descripcion
				distrito = getattr(getattr(persona, 'personaindividual_set', None).first(), 'dis_id', None)
				zona = getattr(getattr(persona, 'personaindividual_set', None).first(), 'zon_id', None)
				writer.writerow([
					f"{persona.per_run}-{persona.per_dv}",
					persona.per_nombres,
					f"{persona.per_apelpat} {persona.per_apelmat or ''}",
					persona.per_email,
					grupo or '',
					distrito.dis_descripcion if distrito else '',
					zona.zon_descripcion if zona else '',
					getattr(curso, 'cur_descripcion', ''),
					str(valor_cuota),
					str(valor_pagado),
					str(valor_adeudado),
					fecha_ultimo_pago.isoformat() if fecha_ultimo_pago else ''
				])
			return response

		# XLSX export implementation if requested
		if format == 'xlsx':
			try:
				import openpyxl
				from openpyxl.utils import get_column_letter
			except Exception:
				return Response({'detail': 'openpyxl no está instalado en el entorno.'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
			wb = openpyxl.Workbook()
			ws = wb.active
			ws.title = 'Pagos'
			headers = ['Rut', 'Nombre', 'Apellidos', 'Correo electrónico', 'Grupo', 'Distrito', 'Zona', 'Curso', 'Valor Cuota', 'Valor pagado', 'Valor Adeudado', 'Fecha ultimo pago']
			ws.append(headers)
			for p in qs.select_related('per_id', 'cur_id'):
				persona = p.per_id
				curso = getattr(p.cur_id, 'cur_id', p.cur_id) if p.cur_id else None
				valor_cuota = curso.cur_cuota_con_almuerzo if curso and hasattr(curso, 'cur_cuota_con_almuerzo') else ''
				valor_pagado = p.pap_valor
				if valor_cuota and valor_pagado:
					valor_adeudado = Decimal(valor_cuota) - Decimal(valor_pagado)
				else:
					valor_adeudado = ''
				fecha_ultimo_pago = p.pap_fecha_hora
				persona_grupo = PersonaGrupo.objects.filter(per_id=persona).first()
				grupo = persona_grupo.gru_id.gru_descripcion if persona_grupo else ''
				distrito = getattr(getattr(persona, 'personaindividual_set', None).first(), 'dis_id', None)
				zona = getattr(getattr(persona, 'personaindividual_set', None).first(), 'zon_id', None)
				row = [
					f"{persona.per_run}-{persona.per_dv}",
					persona.per_nombres,
					f"{persona.per_apelpat} {persona.per_apelmat or ''}",
					persona.per_email,
					grupo or '',
					distrito.dis_descripcion if distrito else '',
					zona.zon_descripcion if zona else '',
					getattr(curso, 'cur_descripcion', ''),
					str(valor_cuota),
					str(valor_pagado),
					str(valor_adeudado),
					fecha_ultimo_pago.isoformat() if fecha_ultimo_pago else ''
				]
				ws.append(row)

			# Auto fit columns (basic)
			for idx, col in enumerate(ws.columns, 1):
				max_length = 0
				column = get_column_letter(idx)
				for cell in col:
					try:
						if cell.value and len(str(cell.value)) > max_length:
							max_length = len(str(cell.value))
					except Exception:
						pass
				ws.column_dimensions[column].width = max_length + 2

			output = io.BytesIO()
			wb.save(output)
			output.seek(0)
			response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
			response['Content-Disposition'] = 'attachment; filename="pagos_export.xlsx"'
			return response
		# default fallback
		return Response({'detail': 'Formato no soportado.'}, status=status.HTTP_400_BAD_REQUEST)


class ComprobantePagoViewSet(viewsets.ModelViewSet):
	queryset = ComprobantePago.objects.all()
	serializer_class = ComprobantePagoSerializer
	permission_classes = [IsAuthenticated]
	filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
	filterset_fields = ['usu_id', 'pec_id', 'coc_id']
	search_fields = ['cpa_numero']

	@action(detail=False, methods=['post'], url_path='generar')
	def generar_comprobante(self, request):
		"""Genera un comprobante desde una lista de pagos. Payload: {"pago_ids": [1,2], "coc_id": 1, "pec_id": 2, "cpa_fecha": "2024-01-01", "cpa_numero": 123} """
		pago_ids = request.data.get('pago_ids')
		coc_id = request.data.get('coc_id')
		pec_id = request.data.get('pec_id')
		cpa_numero = request.data.get('cpa_numero')
		cpa_fecha = request.data.get('cpa_fecha')

		if not pago_ids:
			return Response({'detail': 'pago_ids requerido'}, status=status.HTTP_400_BAD_REQUEST)
		pagos = PagoPersona.objects.filter(pap_id__in=pago_ids)
		if not pagos.exists():
			return Response({'detail': 'No se encontraron pagos con los ids entregados.'}, status=status.HTTP_404_NOT_FOUND)

		# calcular valor total
		total = sum([p.pap_valor for p in pagos])

		# Determine usu
		usu = None
		if hasattr(request.user, 'usuario'):
			usu = request.user.usuario
		else:
			from usuarios.models import Usuario as UsuarioModel
			usu = UsuarioModel.objects.filter(usu_username=getattr(request.user, 'username', None)).first() if getattr(request.user, 'username', None) else None
			if not usu and getattr(request.user, 'email', None):
				usu = UsuarioModel.objects.filter(usu_email=request.user.email).first()
		if not usu:
			return Response({'detail': 'No se pudo determinar Usuario (usu_id).'}, status=status.HTTP_400_BAD_REQUEST)

		comprobante_data = {
			'usu_id': usu,
			'pec_id': pec_id,
			'coc_id': coc_id,
			'cpa_fecha_hora': timezone.now(),
			'cpa_fecha': cpa_fecha or timezone.now().date(),
			'cpa_numero': cpa_numero or 0,
			'cpa_valor': total,
		}
		serializer = self.get_serializer(data=comprobante_data)
		serializer.is_valid(raise_exception=True)
		comprobante = serializer.save()

		# vincular pagos al comprobante
		for p in pagos:
			PagoComprobante.objects.create(pap_id=p, cpa_id=comprobante)

		return Response(self.get_serializer(comprobante).data, status=status.HTTP_201_CREATED)


class PagoComprobanteViewSet(viewsets.ModelViewSet):
	queryset = PagoComprobante.objects.all()
	serializer_class = PagoComprobanteSerializer
	permission_classes = [IsAuthenticated]


class PagoCambioPersonaViewSet(viewsets.ModelViewSet):
	queryset = PagoCambioPersona.objects.all()
	serializer_class = PagoCambioPersonaSerializer
	permission_classes = [IsAuthenticated]


class PrepagoViewSet(viewsets.ModelViewSet):
	queryset = Prepago.objects.all()
	serializer_class = PrepagoSerializer
	permission_classes = [IsAuthenticated]


